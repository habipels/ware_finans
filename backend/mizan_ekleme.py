
import openpyxl
import sqlite3
baglanti  = sqlite3.connect(r"C:\Users\habip\Documents\GitHub\ware_finans\backend\db.sqlite3")

cursor  = baglanti.cursor()
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r'C:\Users\habip\Documents\GitHub\ware_finans\backend\mizan.xlsx')
# read by default 1st sheet of an excel file
def veri_al(id,hesap_kodu,hesap_adi):
    cursor.execute("insert into bilgideposu_hesapplanlari Values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(id,hesap_kodu,hesap_adi,"Hayır",None,None,0,0,0,None,None,"",None,"",hesap_adi,None,"","",0,"","","","","","","","","","","","",0,None,"","",1,0,None,None,None))
    baglanti.commit()
 
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    a= []
    a.append(row+356)
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        a.append(col[row].value,)
    print(a)
    veri_al(a[0],a[1],a[2])
