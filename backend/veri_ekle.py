
import openpyxl
import sqlite3
baglanti  = sqlite3.connect(r"C:\Users\habip\Documents\GitHub\ware_finans\backend\db.sqlite3")

cursor  = baglanti.cursor()
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r'C:\Users\habip\Documents\GitHub\ware_finans\backend\de.xlsx')
# read by default 1st sheet of an excel file
def veri_al(id,N1,N2,N3,N4,amortismana_tabi_olan_ikstisidai_kiymet,faydali_omur,amortisman_orani,defter_beyan_kodu):
    if faydali_omur:
        pass
    else:
        faydali_omur = 0
    if amortisman_orani:
        pass
    else:
        amortisman_orani = 0
    cursor.execute("insert into bilgideposu_amortisman_bilgileri Values(?,?,?,?,?,?,?,?,?)",(id,N3,N4,amortismana_tabi_olan_ikstisidai_kiymet,faydali_omur,amortisman_orani,defter_beyan_kodu,N1,N2))
    baglanti.commit()
 
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    a= []
    a.append(row+1)
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        a.append(col[row].value,)
    print(a)
    veri_al(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7],a[8])

#CREATE TABLE "bilgideposu_amortisman_bilgileri" ("id" integer NOT NULL PRIMARY KEY AUTOINCREMENT, "N3" varchar(200) NULL, "N4" varchar(200) NULL, "amortismana_tabi_olan_ikstisidai_kiymet" varchar(400) NULL, "faydali_omur" bigint NOT NULL, "amortisman_orani" real NOT NULL, "defter_beyan_kodu" varchar(200) NULL, "N1" varchar(200) NULL, "N2" varchar(200) NULL)