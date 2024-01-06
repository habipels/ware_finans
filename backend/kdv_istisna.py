
import openpyxl
import sqlite3
baglanti  = sqlite3.connect(r"C:\Users\habip\Documents\GitHub\ware_finans\backend\db.sqlite3")

cursor  = baglanti.cursor()
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r'C:\Users\habip\Documents\GitHub\ware_finans\backend\kdvistisna.xlsx')
# read by default 1st sheet of an excel file
def veri_al(id,kod,icerik):

    cursor.execute("insert into bilgideposu_kdv_istisna_kodu Values(?,?,?,?)",(id,kod,icerik,""))
    baglanti.commit()
 
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    a= []
    a.append(row+4)
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        a.append(col[row].value,)
    print(a)
    veri_al(a[0],a[1],a[2])

#CREATE TABLE "bilgideposu_amortisman_bilgileri" ("id" integer NOT NULL PRIMARY KEY AUTOINCREMENT, "N3" varchar(200) NULL, "N4" varchar(200) NULL, "amortismana_tabi_olan_ikstisidai_kiymet" varchar(400) NULL, "faydali_omur" bigint NOT NULL, "amortisman_orani" real NOT NULL, "defter_beyan_kodu" varchar(200) NULL, "N1" varchar(200) NULL, "N2" varchar(200) NULL)